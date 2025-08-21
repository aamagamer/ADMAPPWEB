from flask import Flask, request, jsonify, send_from_directory, render_template
import pyodbc
from flask_cors import CORS
from datetime import datetime, date, time
from flask import Flask, session, redirect, url_for
from flask import send_file
import pandas as pd
from dotenv import load_dotenv
import os
from io import BytesIO
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from apscheduler.schedulers.background import BackgroundScheduler
from functools import wraps
from decimal import Decimal

app = Flask(__name__, static_url_path='', static_folder='.')
CORS(app)

app.secret_key = 'clave_secreta_segura'

load_dotenv()




def get_connection():
    return pyodbc.connect(
        f"DRIVER={os.getenv('DB_DRIVER')};"
        f"SERVER={os.getenv('DB_SERVER')};"
        f"DATABASE={os.getenv('DB_NAME')};"
        f"UID={os.getenv('DB_USER')};"
        f"PWD={os.getenv('DB_PASSWORD')};"
        f"Trusted_Connection={os.getenv('DB_TRUSTED_CONNECTION')};"
        f"Encrypt={os.getenv('DB_ENCRYPT')};"
        f"Connection Timeout={os.getenv('DB_TIMEOUT')};"
    )


def requiere_rol(rol_permitido):
    def decorador(f):
        @wraps(f)
        def funcion_envuelta(*args, **kwargs):
            if 'id_usuario' not in session or session.get('rol') != rol_permitido:
                return redirect(url_for('index'))  
            return f(*args, **kwargs)
        return funcion_envuelta
    return decorador

def calcular_dias_vacaciones(fecha_ingreso_str):
    hoy = datetime.now().date()
    fecha_ingreso = datetime.strptime(fecha_ingreso_str, '%Y-%m-%d').date()

    años = hoy.year - fecha_ingreso.year
    if (hoy.month, hoy.day) < (fecha_ingreso.month, fecha_ingreso.day):
        años -= 1

    if años < 1:
        return 0
    elif años == 1:
        return 12
    elif años == 2:
        return 14
    elif años == 3:
        return 16
    elif años == 4:
        return 18
    elif años == 5:
        return 20
    elif 6 <= años <= 10:
        return 22
    elif 11 <= años <= 15:
        return 24
    elif 16 <= años <= 20:
        return 26
    elif 21 <= años <= 25:
        return 28
    else:
        return 30

@app.route('/api/usuario/<int:id>/area', methods=['GET'])
def obtener_areas_usuario(id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT a.idArea, a.NombreArea
        FROM Usuario_Area ua
        JOIN Area a ON ua.idArea = a.idArea
        WHERE ua.idUsuario = ?
    """, (id,))
    rows = cursor.fetchall()
    conn.close()

    areas = [{"idArea": row[0], "NombreArea": row[1]} for row in rows]
    return jsonify(areas)

    
    


@app.route('/api/solicitarVacaciones', methods=['POST'])
def solicitar_vacaciones():
    conn = None

    try:
        data = request.json
        id_usuario = data.get('idUsuario')
        fecha_salida = data.get('fechaInicio')
        fecha_regreso = data.get('fechaFin')
        motivo = data.get('motivo')
        dias_solicitados = data.get('diasSolicitados')

        if not all([id_usuario, fecha_salida, fecha_regreso, dias_solicitados]):
            return jsonify({"error": "Faltan datos requeridos"}), 400

        try:
            dias_solicitados = int(dias_solicitados)
        except (ValueError, TypeError):
            return jsonify({"error": "Días solicitados inválidos"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Verificar existencia del usuario y días disponibles
        cursor.execute("SELECT DiasDisponibles FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Usuario no encontrado"}), 404

        dias_disponibles = row[0] or 0

        if dias_solicitados > dias_disponibles:
            return jsonify({
                "error": f"No tienes suficientes días disponibles. Disponibles: {dias_disponibles}, Solicitados: {dias_solicitados}"
            }), 400

        # Obtener el TipoRol del usuario
        cursor.execute("""
            SELECT r.idRol FROM Usuario u
            JOIN Rol r ON u.Rol_idRol = r.idRol
            WHERE u.idUsuario = ?
        """, (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Rol del usuario no encontrado"}), 400

        tipo_rol = rol_row[0]

        # Asignar estado_id según el tipo de rol
        if tipo_rol == 1:  # Empleado
            estado_id = 18  
        elif tipo_rol == 4:  # Líder de Área
            estado_id = 19
        elif tipo_rol == 3:  # RH
            estado_id = 20
        else:
            return jsonify({"error": f"Rol no válido o no autorizado: {tipo_rol}"}), 400

        # Insertar la solicitud de vacaciones
        cursor.execute("""
            INSERT INTO Vacaciones (
                Usuario_idUsuario,
                EstadoSolicitud_idSolicitud,
                DiasSolicitados,
                FechaSalida,
                FechaRegreso
            ) VALUES (?, ?, ?, ?, ?)
        """, (id_usuario, estado_id, dias_solicitados, fecha_salida, fecha_regreso))

        # Actualizar los días disponibles del usuario
        cursor.execute("""
            UPDATE Usuario
            SET DiasDisponibles = DiasDisponibles - ?
            WHERE idUsuario = ?
        """, (dias_solicitados, id_usuario))

        conn.commit()

        return jsonify({
            "mensaje": "Solicitud registrada correctamente",
            "dias_solicitados": dias_solicitados,
            "estadoInicial_id": estado_id
        }), 201

    except Exception as e:
        print("ERROR EN SOLICITAR VACACIONES:", e)
        return jsonify({"error": "Error al registrar solicitud"}), 500

    finally:
        if conn:
            conn.close()

    
@app.route('/api/diafestivo', methods=['POST'])
def insertar_dia_festivo():
    data = request.get_json()

    fecha = data.get('fecha')
    descripcion = data.get('descripcion')
    anio = data.get('anio')

    if not fecha or not descripcion or not anio:
        return jsonify({'error': 'Todos los campos son obligatorios'}), 400

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO DiasFestivos (Fecha, Descripcion, Anio)
            VALUES (?, ?, ?)
        """, (fecha, descripcion, anio))
        conn.commit()
        return jsonify({'mensaje': 'Día festivo insertado correctamente'}), 201

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()


@app.route('/api/diasfestivos', methods=['GET'])
def obtener_dias_festivos():
    anio = request.args.get('anio', default=datetime.now().year, type=int)

    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT fecha, descripcion
            FROM DiasFestivos
            WHERE anio = ?
        """, anio)

        rows = cursor.fetchall()
        resultado = [
            {"fecha": row.fecha.strftime('%Y-%m-%d'), "descripcion": row.descripcion}
            for row in rows
        ]

        return jsonify(resultado), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()




@app.route('/api/agregarArea', methods=['POST'])
def insertar_area():
    data = request.get_json()

    nombreArea = data.get('Area')

    if not nombreArea:
        return jsonify({'error': 'Faltan datos'}), 400

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Area (NombreArea) VALUES (?)", (nombreArea,))
        conn.commit()
        return jsonify({'mensaje': 'Área ingresada correctamente'}), 201

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()

@app.route('/api/agregarAcceso', methods=['POST'])
def insertar_acceso():
    data = request.get_json()  # Recibimos los datos en formato JSON desde el cliente

    nombreAcceso = data.get('Acceso')  # Extraemos el campo 'Acceso' del JSON

    if not nombreAcceso:  # Validamos que no venga vacío
        return jsonify({'error': 'Faltan datos'}), 400

    try:
        conn = get_connection()   # Abrimos conexión a la BD
        cursor = conn.cursor()

        # Query parametrizada para evitar inyección SQL
        cursor.execute("INSERT INTO Acceso (NombreAcceso) VALUES (?)", (nombreAcceso,))
        conn.commit()  # Guardamos los cambios en la BD

        return jsonify({'mensaje': 'Acceso ingresado correctamente'}), 201

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()



@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/login', methods=['POST'])
def login():
    username = request.form.get('username')
    password = request.form.get('password')

    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Verificar usuario y obtener su rol
        cursor.execute("""
            SELECT u.idUsuario, u.clave, r.TipoRol
            FROM Usuario u
            JOIN Rol r ON u.Rol_idRol = r.idRol
            WHERE u.idUsuario = ?
        """, (username,))
        row = cursor.fetchone()

        if row and row[1] == password:
            # Obtener FechaIngreso y vacaciones actuales
            cursor.execute("SELECT FechaIngreso, Vacaciones FROM Usuario WHERE idUsuario = ?", (username,))
            ingreso_row = cursor.fetchone()

            if ingreso_row:
                fecha_ingreso = ingreso_row[0]
                vacaciones_actuales = ingreso_row[1] or 0
                fecha_ingreso_str = fecha_ingreso.strftime('%Y-%m-%d')
                vacaciones_calculadas = calcular_dias_vacaciones(fecha_ingreso_str)

                if vacaciones_actuales != vacaciones_calculadas:
                    cursor.execute(
                        "UPDATE Usuario SET Vacaciones = ? WHERE idUsuario = ?",
                        (vacaciones_calculadas, username)
                    )
                    conn.commit()

            # Guardar en sesión
            rol_normalizado = row[2].strip().lower()
            session['id_usuario'] = row[0]
            session['rol'] = rol_normalizado

            rutas = {
                'empleado': '/empleado',
                'rh': '/rh',
                'administrador': '/admin',
                'lider area': '/lider'
            }

            return jsonify({
                "success": True,
                "redirect": rutas.get(rol_normalizado, "/"),
                "idUsuario": row[0]
            })

        return jsonify({"success": False, "message": "Usuario o contraseña incorrectos"})

    except Exception as e:
        return jsonify({"success": False, "message": f"Error al conectar: {e}"})

    finally:
        conn.close()


# Rutas protegidas según el rol del usuario

@app.route('/admin')
@requiere_rol('administrador')
def admin():
    return send_from_directory('.', 'admin.html')

@app.route('/empleado')
@requiere_rol('empleado')
def empleado():
    return send_from_directory('.', 'empleado.html')

@app.route('/rh')
@requiere_rol('rh')
def rh():
    return send_from_directory('.', 'rh.html')

@app.route('/lider')
@requiere_rol('lider area')
def lider():
    return send_from_directory('.', 'lider.html')



@app.route('/api/empleados', methods=['GET'])
def obtener_empleados():
    estado = request.args.get('estado')  # puede ser 'activos', 'inactivos' o 'todos'
    area_id = request.args.get('area')   # ID del área a filtrar
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Query base - SIEMPRE incluir el JOIN con Usuario_Area si queremos filtrar por área
    if area_id:
        base_query = """
            SELECT DISTINCT Usuario.idUsuario, nombres, paterno, materno, puesto, Rol_idRol, Usuario.Estado
            FROM Usuario
            INNER JOIN Usuario_Area ON Usuario.idUsuario = Usuario_Area.idUsuario
        """
    else:
        base_query = """
            SELECT Usuario.idUsuario, nombres, paterno, materno, puesto, Rol_idRol, Usuario.Estado
            FROM Usuario
        """
    
    # Condiciones WHERE
    conditions = []
    params = []
    
    # Si se especifica un área, agregar condición
    if area_id:
        conditions.append("Usuario_Area.idArea = ?")
        params.append(area_id)
    
    # Filtro por estado - IMPORTANTE: manejar correctamente el caso "todos"
    if estado == 'activos':
        conditions.append("Usuario.Estado = 'Activo'")
    elif estado == 'inactivos':
        conditions.append("Usuario.Estado = 'Inactivo'")
    # Si es 'todos', no agregamos condición de estado (mostramos activos e inactivos)
    
    # Construir query final
    if conditions:
        query = base_query + " WHERE " + " AND ".join(conditions)
    else:
        query = base_query
    
    #print(f"Query ejecutada: {query}")  # Para debug
    #print(f"Parámetros: {params}")      # Para debug
    
    cursor.execute(query, params)
    empleados = [
        {
            "id": row.idUsuario,
            "nombre": f"{row.nombres} {row.paterno} {row.materno}",
            "puesto": row.puesto,
            "Rol_idRol": row.Rol_idRol,
            "estado": row.Estado  # Incluir estado para debug
        } for row in cursor.fetchall()
    ]
    
    conn.close()
    return jsonify(empleados)

@app.route('/api/actualizarEstadoSolicitud', methods=['POST'])
def actualizar_estado_solicitud():
    try:
        data = request.json
        id_vacacion = data.get("idVacaciones")
        accion = data.get("accion")  # 'aceptar' o 'rechazar'
        id_usuario = data.get("idUsuario")  # ID del usuario que hace la acción

        if not all([id_vacacion, accion, id_usuario]):
            return jsonify({"error": "Faltan datos"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Obtener rol del usuario que actualiza
        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Usuario no encontrado"}), 404
        tipo_rol = rol_row[0]

        # Obtener estado actual de la solicitud
        cursor.execute("SELECT EstadoSolicitud_idSolicitud FROM Vacaciones WHERE idVacaciones = ?", (id_vacacion,))
        estado_actual = cursor.fetchone()
        if not estado_actual:
            return jsonify({"error": "Solicitud no encontrada"}), 404
        estado_actual = estado_actual[0]

        # Decidir nuevo estado según rol y acción
        nuevo_estado = None

        if accion == "rechazar":
            nuevo_estado = 1  # Rechazado

        elif accion == "aceptar":
            if tipo_rol == 4:       # Líder
                nuevo_estado = 19   # Pendiente RH
            elif tipo_rol == 3:     # RH
                nuevo_estado = 2   # Aceptado
            elif tipo_rol == 2:     # Admin
                nuevo_estado = 2    # Aceptado
            else:
                return jsonify({"error": "Rol no autorizado para aceptar"}), 403

        else:
            return jsonify({"error": "Acción inválida"}), 400

        # Actualizar estado
        cursor.execute("""
            UPDATE Vacaciones
            SET EstadoSolicitud_idSolicitud = ?
            WHERE idVacaciones = ?
        """, (nuevo_estado, id_vacacion))

        # Si rechazó, devolver días disponibles
        if nuevo_estado == 1:
            cursor.execute("""
                SELECT Usuario_idUsuario, DiasSolicitados
                FROM Vacaciones
                WHERE idVacaciones = ?
            """, (id_vacacion,))
            res = cursor.fetchone()
            if res:
                id_usuario_solicitud, dias_solicitados = res
                if dias_solicitados > 0:
                    cursor.execute("""
                        UPDATE Usuario
                        SET DiasDisponibles = DiasDisponibles + ?
                        WHERE idUsuario = ?
                    """, (dias_solicitados, id_usuario_solicitud))

        conn.commit()
        return jsonify({"mensaje": "Estado actualizado", "nuevo_estado": nuevo_estado}), 200

    except Exception as e:
        print("Error:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if conn:
            conn.close()

@app.route('/api/actualizarEstadoPermiso', methods=['POST'])
def actualizar_estado_permiso():
    conn = None
    try:
        data = request.json
        id_permiso = data.get("idPermiso")
        accion = data.get("accion")  # 'aceptar' o 'rechazar'
        id_usuario = data.get("idUsuario")  # ID del usuario que hace la acción

        if not all([id_permiso, accion, id_usuario]):
            return jsonify({"error": "Faltan datos"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Obtener rol del usuario que realiza la acción
        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Usuario no encontrado"}), 404
        tipo_rol = rol_row[0]

        # Obtener estado actual del permiso (opcional, pero lo dejamos por consistencia)
        cursor.execute("SELECT EstadoSolicitud_idSolicitud FROM Permiso WHERE idPermiso = ?", (id_permiso,))
        estado_actual = cursor.fetchone()
        if not estado_actual:
            return jsonify({"error": "Permiso no encontrado"}), 404
        estado_actual = estado_actual[0]

        # Determinar el nuevo estado
        if accion == "rechazar":
            nuevo_estado = 1  # Rechazado
        elif accion == "aceptar":
            if tipo_rol == 4:       # Líder de Área
                nuevo_estado = 19   # Pendiente RH
            elif tipo_rol in (3, 2):  # RH o Admin
                nuevo_estado = 2    # Aceptado
            else:
                return jsonify({"error": "Rol no autorizado para aceptar"}), 403
        else:
            return jsonify({"error": "Acción inválida"}), 400

        # Actualizar el estado de la solicitud
        cursor.execute("""
            UPDATE Permiso
            SET EstadoSolicitud_idSolicitud = ?
            WHERE idPermiso = ?
        """, (nuevo_estado, id_permiso))

        conn.commit()
        return jsonify({
            "mensaje": "Estado actualizado correctamente",
            "nuevo_estado": nuevo_estado
        }), 200

    except Exception as e:
        print("❌ ERROR:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if conn:
            conn.close()


@app.route('/api/empleado/<int:id>', methods=['GET'])
def obtener_empleado(id):
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Consulta principal del usuario
        cursor.execute("""
            SELECT u.*, r.TipoRol, a.NombreAcceso
            FROM Usuario u
            LEFT JOIN Rol r ON u.Rol_idRol = r.idRol
            LEFT JOIN Acceso a on u.Acceso_idAcceso = a.idAcceso
            WHERE u.idUsuario = ?
        """, id)

        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Empleado no encontrado"}), 404

        columns = [col[0] for col in cursor.description]
        usuario = dict(zip(columns, row))

        # 👉 Formatear las fechas: solo día-mes-año
        for key, value in usuario.items():
            if isinstance(value, (date, datetime)):
                usuario[key] = value.strftime("%Y-%m-%d")  # o "%d/%m/%Y" si prefieres ese formato

        # Consulta de áreas

        cursor.execute("""
            SELECT a.idArea, a.NombreArea
            FROM Usuario_Area ua
            JOIN Area a ON ua.idArea = a.idArea
            WHERE ua.idUsuario = ?
        """, id)

        usuario["Areas"] = [{"idArea": r[0], "NombreArea": r[1]} for r in cursor.fetchall()]

        return jsonify(usuario)

    except Exception as e:
        print(f"Error al obtener empleado: {e}")
        return jsonify({"error": "Error interno del servidor"}), 500
    finally:
        conn.close()



@app.route('/api/empleado/<int:id>', methods=['PUT'])
def actualizar_empleado(id):
    data = request.get_json()
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Verifica que el usuario exista
        cursor.execute("SELECT 1 FROM Usuario WHERE idUsuario = ?", id)
        if not cursor.fetchone():
            return jsonify({"error": "Empleado no encontrado"}), 404

        # Actualizar datos del usuario (sin Area_idArea porque ya no existe)
        cursor.execute("""
            UPDATE Usuario SET
                Rol_idRol = ?, Nombres = ?, Paterno = ?, Materno = ?,
                FechaNacimiento = ?, Direccion = ?, CodigoPostal = ?, Correo = ?, NSS = ?, Telefono = ?,
                FechaIngreso = ?, RFC = ?, Curp = ?, Puesto = ?, NombreContactoEmergencia = ?,
                TelefonoEmergencia = ?, Parentesco = ?, clave = ?,
                SueldoDiario = ?, SueldoSemanal = ?, BonoSemanal = ?, Mensual = ?, Vacaciones = ?, DiasDisponibles = ?,
                Acceso_idAcceso = ?, NumeroAcceso = ?
            WHERE idUsuario = ?
        """, (
            data['rol_id'],
            data['nombres'],
            data['paterno'],
            data['materno'],
            data['fechaNacimiento'],
            data['direccion'],
            data['codigoPostal'],
            data['correo'],
            data['nss'],
            data['telefono'],
            data['fechaIngreso'],
            data['rfc'],
            data['curp'],
            data['puesto'],
            data['nombreContactoEmergencia'],
            data['telefonoEmergencia'],
            data['parentesco'],
            data['contraseña'],
            float(data['SueldoDiario']),
            float(data['SueldoSemanal']),
            float(data['BonoSemanal']),
            float(data['Mensual']),
            int(data['Vacaciones']),
            int(data['diasDisponibles']),
            data['Acceso_idAcceso'],
            data['NumeroAcceso'],
            id
        ))

        # 1. Eliminar áreas anteriores del usuario
        cursor.execute("DELETE FROM Usuario_Area WHERE idUsuario = ?", id)

        # 2. Insertar las nuevas áreas (si hay)
        areas = data.get("areas", [])
        for area_id in areas:
            cursor.execute(
                "INSERT INTO Usuario_Area (idUsuario, idArea) VALUES (?, ?)",
                (id, area_id)
            )

        conn.commit()
        return jsonify({"mensaje": "Empleado actualizado correctamente"}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route('/api/empleado', methods=['POST'])
def agregar_empleado():
    data = request.get_json()
    try:
        # Validaciones básicas
        required_fields = ['idUsuario', 'rol_id', 'nombres', 'paterno', 'fechaNacimiento', 
                          'direccion', 'codigoPostal', 'correo', 'nss', 'telefono', 
                          'fechaIngreso', 'rfc', 'curp', 'puesto', 'nombreContactoEmergencia',
                          'telefonoEmergencia', 'parentesco', 'contraseña']
        
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({"error": f"Campo requerido: {field}"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # 🔧 CORRECCIÓN: Los nombres de los parámetros deben coincidir
        # Frontend envía: "Empresa" (ID de la empresa) y "Acceso" (número de acceso)
        # Backend necesita: Acceso_idAcceso y NombreAcceso (o NumeroAcceso)
        
        # Manejar los campos opcionales de empresa/acceso
        acceso_id = None
        numero_acceso = None
        
        # Si se proporcionó una empresa
        if data.get('Empresa') and data['Empresa'] != '':
            acceso_id = int(data['Empresa'])  # Este es el ID de la tabla Acceso
        
        # Si se proporcionó número de acceso
        if data.get('Acceso') and data['Acceso'].strip() != '':
            numero_acceso = data['Acceso'].strip()

        # Insertar en Usuario
        cursor.execute("""
            INSERT INTO Usuario (
                idUsuario, Rol_idRol, Nombres, Paterno, Materno,
                FechaNacimiento, Direccion, CodigoPostal, Correo, NSS, Telefono,
                FechaIngreso, RFC, Curp, Puesto, NombreContactoEmergencia,
                TelefonoEmergencia, Parentesco, FechaBaja, ComentarioSalida,
                clave, Estado, SueldoDiario, SueldoSemanal, BonoSemanal, Mensual, 
                Vacaciones, DiasDisponibles, Acceso_idAcceso, NumeroAcceso
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, NULL,
                    ?, 'Activo', ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            int(data['idUsuario']),
            int(data['rol_id']),
            data['nombres'],
            data['paterno'],
            data.get('materno', ''),  # Puede ser vacío
            data['fechaNacimiento'],
            data['direccion'],
            data['codigoPostal'],
            data['correo'],
            data['nss'],
            data['telefono'],
            data['fechaIngreso'],
            data['rfc'],
            data['curp'],
            data['puesto'],
            data['nombreContactoEmergencia'],
            data['telefonoEmergencia'],
            data['parentesco'],
            data['contraseña'],
            float(data.get('SueldoDiario', 0)),
            float(data.get('SueldoSemanal', 0)),
            float(data.get('BonoSemanal', 0)),
            float(data.get('Mensual', 0)),
            int(data.get('Vacaciones', 0)),
            int(data.get('diasDisponibles', 0)),
            acceso_id,        # 🔧 Puede ser None
            numero_acceso     # 🔧 Puede ser None
        ))

        # Manejar áreas
        area_ids = data.get('areas')
        if not area_ids or not isinstance(area_ids, list):
            return jsonify({"error": "Debes proporcionar al menos una área (lista de IDs)"}), 400

        # Insertar en tabla intermedia Usuario_Area
        for id_area in area_ids:
            # Verificar que el área existe
            cursor.execute("SELECT 1 FROM Area WHERE idArea = ?", (int(id_area),))
            if not cursor.fetchone():
                return jsonify({"error": f"Área con ID {id_area} no existe"}), 400
            
            # Insertar relación
            cursor.execute("INSERT INTO Usuario_Area (idUsuario, idArea) VALUES (?, ?)",
                           (int(data['idUsuario']), int(id_area)))

        conn.commit()
        return jsonify({"mensaje": "Empleado insertado correctamente"}), 201

    except sqlite3.IntegrityError as e:
        return jsonify({"error": f"Error de integridad: {str(e)}"}), 400
    except ValueError as e:
        return jsonify({"error": f"Error en formato de datos: {str(e)}"}), 400
    except Exception as e:
        print(f"Error inesperado: {str(e)}")  # Para debug
        return jsonify({"error": f"Error interno: {str(e)}"}), 500
    finally:
        if 'conn' in locals():
            conn.close()



@app.route('/api/empleado/baja/<int:id>', methods=['PUT'])
def baja_empleado(id):
    conn = None
    try:
        data = request.json
        fecha_baja = data.get("fechaBaja")
        comentario = data.get("comentarioSalida")
        id_razon_baja = data.get("idRazonBaja")  # <-- NUEVO

        if not fecha_baja or not comentario or not id_razon_baja:
            return jsonify({"error": "Todos los campos son obligatorios"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT 1 FROM Usuario WHERE idUsuario = ?", (id,))
        if not cursor.fetchone():
            return jsonify({"error": "Empleado no encontrado"}), 404

        cursor.execute("""
            UPDATE Usuario
            SET Estado = 'Inactivo',
                FechaBaja = ?,
                ComentarioSalida = ?,
                idRazonBaja = ?  -- <-- NUEVO
            WHERE idUsuario = ?
        """, (fecha_baja, comentario, id_razon_baja, id))

        conn.commit()
        return jsonify({"mensaje": "Empleado dado de baja correctamente"}), 200

    except Exception as e:
        print("Error al dar de baja:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if conn:
            conn.close()

@app.route('/api/usuario/activar/<int:idUsuario>', methods=['PUT'])
def activar_usuario(idUsuario):
    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE Usuario
            SET Estado = 'Activo'
            WHERE idUsuario = ?
        """, (idUsuario,))
        conn.commit()

        # Verificamos si realmente se actualizó alguna fila
        if cursor.rowcount == 0:
            return jsonify({'error': f'No se encontró usuario con id {idUsuario}'}), 404

        return jsonify({'mensaje': f'Usuario {idUsuario} activado correctamente'}), 200

    except Exception as e:
        print(f'❌ ERROR EN /api/usuario/activar/{idUsuario}: {e}')
        return jsonify({'error': 'Error al activar el usuario'}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/acceso', methods=['GET'])
def obtener_accesos():
    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT idAcceso, NombreAcceso
            FROM Acceso
        """)
        rows = cursor.fetchall()

        accesos = [{'idAcceso': row[0], 'NombreAcceso': row[1]} for row in rows]

        return jsonify(accesos), 200

    except Exception as e:
        print(f'❌ ERROR EN /api/acceso: {e}')
        return jsonify({'error': 'Error al obtener accesos'}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



@app.route('/api/razones-baja', methods=['GET'])
def obtener_razones_baja():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT idRazonBaja, RazonBaja FROM RazonesBaja")
    
    razones = [{"idRazonBaja": row[0], "RazonBaja": row[1]} for row in cursor.fetchall()]
    
    conn.close()
    return jsonify(razones)

@app.route('/api/usuario/<int:idUsuario>/reportes/count', methods=['GET'])
def contar_reportes_usuario(idUsuario):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT COUNT(*) 
            FROM Reporte 
            WHERE Usuario_idUsuario = ?
        """, (idUsuario,))
        
        total = cursor.fetchone()[0]

        return jsonify({
            "idUsuario": idUsuario,
            "totalReportes": total
        }), 200

    except Exception as e:
        print(f'❌ ERROR EN /api/usuario/{idUsuario}/reportes/count: {e}')
        return jsonify({'error': 'Error al obtener el conteo de reportes'}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()





@app.route('/api/roles', methods=['GET'])
def obtener_roles():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT idRol, TipoRol FROM Rol")
    roles = [{"idRol": row[0], "TipoRol": row[1]} for row in cursor.fetchall()]
    conn.close()
    return jsonify(roles)

@app.route('/api/areas', methods=['GET'])
def obtener_areas():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT idArea, NombreArea FROM Area")
    areas = [{"idArea": row[0], "NombreArea": row[1]} for row in cursor.fetchall()]
    conn.close()
    return jsonify(areas)

@app.route('/api/compensaciones', methods=['GET'])
def obtener_compensaciones():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT idCompensacion, TipoCompensacion FROM Compensacion")
        resultados = cursor.fetchall()
        cursor.close()
        conn.close()

        compensaciones = [{"id": row[0], "nombre": row[1]} for row in resultados]
        return jsonify({"compensaciones": compensaciones})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/vacante', methods=['POST'])
def crear_vacante():
    data = request.get_json()

    # Validación básica (puedes mejorarla si quieres)
    required_fields = ['area_idarea', 'Usuario_idUsuario', 'Puesto', 'Perfil', 'Habilidades', 'cantidad']
    if not all(field in data for field in required_fields):
        return jsonify({'error': 'Faltan campos requeridos'}), 400

    try:
        conn = get_connection()
        cursor = conn.cursor()
        query = """
            INSERT INTO Vacante (area_idarea, Usuario_idUsuario, Puesto, Perfil, Habilidades, Estado, cantidad)
            VALUES (?, ?, ?, ?, ?, 'Activo',?)
        """
        cursor.execute(query, (
            data['area_idarea'],
            data['Usuario_idUsuario'],
            data['Puesto'],
            data['Perfil'],
            data['Habilidades'],
            int(data['cantidad']),
        ))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Vacante creada exitosamente'}), 201

    except Exception as e:
        print(f"❌ ERROR en /api/vacante: {e}")
        return jsonify({'error': 'Error al crear la vacante'}), 500

    
    


@app.route('/api/usuario/nombre', methods=['POST'])
def obtener_nombre_usuario():
    data = request.json
    id_usuario = data.get("idUsuario")

    if not id_usuario:
        return jsonify({"error": "Falta el ID del usuario"}), 400

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT nombres, paterno, materno
            FROM Usuario
            WHERE idUsuario = ?
        """, id_usuario)
        row = cursor.fetchone()
        conn.close()

        if row:
            return jsonify({
                "nombres": row[0],
                "paterno": row[1],
                "materno": row[2]
            })
        else:
            return jsonify({"error": "Usuario no encontrado"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/api/festivos', methods=['GET'])
def obtener_festivos():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Fecha FROM DiasFestivos")
        resultados = cursor.fetchall()
        fechas = [row[0].strftime('%Y-%m-%d') for row in resultados]
        return jsonify(fechas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()






@app.route('/api/vacaciones/area/<ids>', methods=['GET'])
def obtener_vacaciones_por_areas(ids):
    conn = None
    cursor = None

    try:
        id_usuario = request.args.get('idUsuario')
        if not id_usuario:
            return jsonify({"error": "Se requiere el idUsuario"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Usuario no encontrado"}), 404

        tipo_rol = rol_row[0]

        lista_ids = ids.split(",")
        placeholders = ",".join("?" for _ in lista_ids)

        if tipo_rol == 4:  # Líder
            estado = 18
            query = f"""
                SELECT DISTINCT v.idVacaciones, v.FechaSalida, v.FechaRegreso, u.Nombres, u.Paterno, u.Materno
                FROM Vacaciones v
                JOIN Usuario u ON v.Usuario_idUsuario = u.idUsuario
                JOIN Usuario_Area ua ON ua.idUsuario = u.idUsuario
                WHERE ua.idArea IN ({placeholders}) 
                  AND v.EstadoSolicitud_idSolicitud = ?
                  AND u.Rol_idRol = 1
            """
            params = lista_ids + [estado]

        elif tipo_rol == 3:  # RH
            estado = 19
            query = """
                SELECT DISTINCT v.idVacaciones, v.FechaSalida, v.FechaRegreso, u.Nombres, u.Paterno, u.Materno
                FROM Vacaciones v
                JOIN Usuario u ON v.Usuario_idUsuario = u.idUsuario
                WHERE v.EstadoSolicitud_idSolicitud = ?
            """
            params = [estado]

        elif tipo_rol == 2:  # Admin
            estados = (18, 19, 20)
            query = """
                SELECT DISTINCT v.idVacaciones, v.FechaSalida, v.FechaRegreso, u.Nombres, u.Paterno, u.Materno
                FROM Vacaciones v
                JOIN Usuario u ON v.Usuario_idUsuario = u.idUsuario
                WHERE v.EstadoSolicitud_idSolicitud IN (?,?,?)
            """
            params = estados

        else:
            return jsonify({"error": "Rol no autorizado para esta operación"}), 403

        cursor.execute(query, params)
        rows = cursor.fetchall()

        vacaciones = [{
            "id": row[0],
            "inicio": row[1].strftime('%Y-%m-%d') if row[1] else "",
            "fin": row[2].strftime('%Y-%m-%d') if row[2] else "",
            "nombre": f"{row[3]} {row[4]} {row[5]}"
        } for row in rows]

        return jsonify(vacaciones)

    except Exception as e:
        print("❌ ERROR EN /api/vacaciones/area/<ids>:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



@app.route('/api/permisos/area/<ids>', methods=['GET'])
def obtener_permisos_por_areas(ids):
    conn = None
    cursor = None

    try:
        id_usuario = request.args.get('idUsuario')
        if not id_usuario:
            return jsonify({"error": "Se requiere el idUsuario"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Usuario no encontrado"}), 404

        tipo_rol = rol_row[0]

        lista_ids = ids.split(",")
        placeholders = ",".join("?" for _ in lista_ids)

        if tipo_rol == 4:  # Líder
            estado = 18
            query = f"""
                SELECT DISTINCT
                    p.idPermiso, 
                    p.DiaSolicitado, 
                    p.HoraInicio, 
                    p.HoraFin,
                    p.Razon,
                    ISNULL(c.TipoCompensacion, 'Sin compensación'),
                    u.Nombres, u.Paterno, u.Materno
                FROM Permiso p
                JOIN Usuario u ON p.Usuario_idUsuario = u.idUsuario
                JOIN Usuario_Area ua ON ua.idUsuario = u.idUsuario
                LEFT JOIN Compensacion c ON p.Compensacion_idCompensacion = c.idCompensacion
                WHERE ua.idArea IN ({placeholders})
                  AND p.EstadoSolicitud_idSolicitud = ?
                  AND u.Rol_idRol = 1
            """
            params = lista_ids + [estado]

        elif tipo_rol == 3:  # RH
            estado = 19
            query = """
                SELECT DISTINCT
                    p.idPermiso, 
                    p.DiaSolicitado, 
                    p.HoraInicio, 
                    p.HoraFin,
                    p.Razon,
                    ISNULL(c.TipoCompensacion, 'Sin compensación'),
                    u.Nombres, u.Paterno, u.Materno
                FROM Permiso p
                JOIN Usuario u ON p.Usuario_idUsuario = u.idUsuario
                LEFT JOIN Compensacion c ON p.Compensacion_idCompensacion = c.idCompensacion
                WHERE p.EstadoSolicitud_idSolicitud = ?
            """
            params = [estado]

        elif tipo_rol == 2:  # Admin
            estados = (18, 19, 20)
            query = """
                SELECT DISTINCT
                    p.idPermiso, 
                    p.DiaSolicitado, 
                    p.HoraInicio, 
                    p.HoraFin,
                    p.Razon,
                    ISNULL(c.TipoCompensacion, 'Sin compensación'),
                    u.Nombres, u.Paterno, u.Materno
                FROM Permiso p
                JOIN Usuario u ON p.Usuario_idUsuario = u.idUsuario
                LEFT JOIN Compensacion c ON p.Compensacion_idCompensacion = c.idCompensacion
                WHERE p.EstadoSolicitud_idSolicitud IN (?,?,?)
            """
            params = estados

        else:
            return jsonify({"error": "Rol no autorizado para esta operación"}), 403

        cursor.execute(query, params)
        rows = cursor.fetchall()

        permisos = [{
            "id": row[0],
            "fecha": row[1].strftime('%Y-%m-%d') if row[1] else "",
            "inicio": row[2].strftime('%H:%M') if row[2] else "",
            "fin": row[3].strftime('%H:%M') if row[3] else "",
            "razon": row[4],
            "compensacion": row[5],
            "nombre": f"{row[6]} {row[7]} {row[8]}"
        } for row in rows]

        return jsonify(permisos)

    except Exception as e:
        print("❌ ERROR EN /api/permisos/area/<ids>:", str(e))
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/solicitud/<tipo>/<id>/usuario', methods=['GET'])
def obtener_usuario_por_solicitud(tipo, id):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        if tipo == 'vacacion':
            cursor.execute("""
                SELECT u.idUsuario, u.Nombres, u.Telefono
                FROM Vacaciones v
                JOIN Usuario u ON v.Usuario_idUsuario = u.idUsuario
                WHERE v.idVacaciones = ?
            """, (id,))
        elif tipo == 'permiso':
            cursor.execute("""
                SELECT u.idUsuario, u.Nombres, u.Telefono
                FROM Permiso p
                JOIN Usuario u ON p.Usuario_idUsuario = u.idUsuario
                WHERE p.idPermiso = ?
            """, (id,))
        else:
            return jsonify({"error": "Tipo de solicitud no válido"}), 400

        result = cursor.fetchone()
        if result:
            return jsonify({
                "idUsuario": result[0],
                "nombre": result[1],
                "telefono": result[2]
            })
        else:
            return jsonify({"error": "Solicitud no encontrada"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn: conn.close()




@app.route('/api/solicitarPermiso', methods=['POST'])
def solicitar_permiso():
    conn = None
    try:
        data = request.get_json()
        print("📥 JSON recibido:", data)

        id_usuario = data.get("idUsuario")
        fecha = data.get("fecha")
        hora_inicio = data.get("horaInicio")
        hora_fin = data.get("horaFin")
        razon = data.get("razon")
        id_compensacion = data.get("idCompensacion")

        if not all([id_usuario, fecha, hora_inicio, hora_fin, razon, id_compensacion]):
            print("❌ Campos faltantes")
            return jsonify({"error": "Faltan campos obligatorios"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Obtener el TipoRol del usuario
        cursor.execute("""
            SELECT r.idRol FROM Usuario u
            JOIN Rol r ON u.Rol_idRol = r.idRol
            WHERE u.idUsuario = ?
        """, (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Rol del usuario no encontrado"}), 400

        tipo_rol = rol_row[0]
        print("🧑 Rol del usuario:", tipo_rol)

        # Asignar estado_id según el tipo de rol
        if tipo_rol == 1:  # Empleado
            estado_id = 18
        elif tipo_rol == 4:  # Líder de Área
            estado_id = 19
        elif tipo_rol == 3:  # RH
            estado_id = 20
        else:
            return jsonify({"error": f"Rol no válido o no autorizado: {tipo_rol}"}), 400

        # INSERT
        print("📤 Ejecutando INSERT con estado_id:", estado_id)
        cursor.execute("""
            INSERT INTO Permiso (
                Usuario_idUsuario, EstadoSolicitud_idSolicitud, DiaSolicitado,
                HoraInicio, HoraFin, Razon, Compensacion_idCompensacion
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (id_usuario, estado_id, fecha, hora_inicio, hora_fin, razon, id_compensacion))

        conn.commit()
        print("✅ Permiso solicitado correctamente")
        return jsonify({
            "mensaje": "Permiso solicitado correctamente",
            "estadoInicial_id": estado_id
        }), 201

    except Exception as e:
        print("❌ ERROR:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if conn:
            conn.close()

@app.route('/api/usuario/solicitudes/<int:id>', methods=['GET'])
def obtener_solicitudes_usuario(id):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # VACACIONES: Traer datos + estado en texto
        cursor.execute("""
    SELECT v.DiasSolicitados, v.FechaSalida, v.FechaRegreso, es.Estado
    FROM Vacaciones v
    JOIN EstadoSolicitud es ON v.EstadoSolicitud_idSolicitud = es.idSolicitud
    WHERE v.Usuario_idUsuario = ?
    ORDER BY v.idVacaciones DESC
""", id)
        vacaciones_rows = cursor.fetchall()
        vacaciones = []
        for row in vacaciones_rows:
            vacaciones.append({
                "estadoSolicitud": row[3],
                "diasSolicitados": row[0],
                "fechaSalida": row[1].strftime('%Y-%m-%d') if row[1] else None,
                "fechaRegreso": row[2].strftime('%Y-%m-%d') if row[2] else None
            })

        # PERMISOS: Traer datos + estado + tipo compensación
        cursor.execute("""
    SELECT p.DiaSolicitado, p.HoraInicio, p.HoraFin, p.Razon,
           es.Estado, c.TipoCompensacion
    FROM Permiso p
    JOIN EstadoSolicitud es ON p.EstadoSolicitud_idSolicitud = es.idSolicitud
    JOIN Compensacion c ON p.Compensacion_idCompensacion = c.idCompensacion
    WHERE p.Usuario_idUsuario = ?
    ORDER BY p.idPermiso DESC
""", id)
        permisos_rows = cursor.fetchall()
        permisos = []
        for row in permisos_rows:
            permisos.append({
                "estadoSolicitud": row[4],
                "diaSolicitado": row[0].strftime('%Y-%m-%d') if row[0] else None,
                "horaInicio": row[1].strftime('%H:%M:%S') if row[1] else None,
                "horaFin": row[2].strftime('%H:%M:%S') if row[2] else None,
                "razon": row[3],
                "tipoCompensacion": row[5]
            })

        return jsonify({
            "vacaciones": vacaciones,
            "permisos": permisos
        })

    except Exception as e:
        print("❌ Error en /api/usuario/solicitudes:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/vacantes', methods=['GET'])
def obtener_vacantes():
    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor()

        query = '''
            SELECT 
                Area.NombreArea, 
                Vacante.Puesto,
                Vacante.cantidad,
                Vacante.Perfil, 
                Vacante.Habilidades,
                Vacante.idVacante
            FROM 
                Area 
            INNER JOIN Vacante ON Area.idArea = Vacante.Area_idArea where 
            Estado = 'Activo'
        '''
        cursor.execute(query)
        rows = cursor.fetchall()

        columnas = [column[0] for column in cursor.description]
        vacantes = [dict(zip(columnas, row)) for row in rows]

        return jsonify({'vacantes': vacantes}), 200

    except Exception as e:
        print(f'❌ ERROR EN /api/vacantes: {e}')
        return jsonify({'error': 'Error al obtener las vacantes'}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/vacantes/<int:id>/restar', methods=['PUT'])
def restar_vacante(id):
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Obtener la cantidad actual
        cursor.execute("SELECT cantidad FROM Vacante WHERE idVacante = ?", (id,))
        result = cursor.fetchone()

        if not result:
            return jsonify({'error': 'Vacante no encontrada'}), 404

        cantidad_actual = result[0]

        if cantidad_actual <= 1:
            return jsonify({'error': 'Cantidad ya es mínima'}), 400

        # Actualizar restando 1
        nueva_cantidad = cantidad_actual - 1
        cursor.execute("UPDATE Vacante SET cantidad = ? WHERE idVacante = ?", (nueva_cantidad, id))
        conn.commit()

        return jsonify({'message': f'Cantidad actualizada a {nueva_cantidad}'}), 200

    except Exception as e:
        print(f"❌ ERROR en PUT /api/vacantes/{id}/restar: {e}")
        return jsonify({'error': 'Error al actualizar la vacante'}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

        
    
@app.route('/api/vacantes/<int:idVacante>', methods=['PUT'])
def actualizar_estado_vacante(idVacante):
    data = request.get_json()
    nuevo_estado = data.get('Estado')

    if not nuevo_estado:
        return jsonify({'error': 'Estado no proporcionado'}), 400

    conn = None
    cursor = None

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE Vacante 
            SET Estado = ? 
            WHERE idVacante = ?
        """, (nuevo_estado, idVacante))
        conn.commit()

        return jsonify({'mensaje': 'Estado de la vacante actualizado correctamente'}), 200

    except Exception as e:
        print(f'❌ ERROR EN /api/vacantes/{idVacante}: {e}')
        return jsonify({'error': 'Error al actualizar la vacante'}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.route('/api/reportes', methods=['POST'])
def crear_reporte():
    conn = None
    cursor = None

    try:
        data = request.get_json()

        id_usuario = data.get('Usuario_idUsuario')
        asunto = data.get('Asunto')
        observaciones = data.get('Observaciones')
        fecha_reporte = data.get('FechaReporte')  # <- nuevo campo

        if not id_usuario or not asunto or not observaciones or not fecha_reporte:
            return jsonify({'error': 'Faltan datos requeridos'}), 400

        conn = get_connection()
        cursor = conn.cursor()

        query = '''
            INSERT INTO Reporte (Usuario_idUsuario, Asunto_idAsunto, Observaciones, FechaReporte)
            VALUES (?, ?, ?, ?)
        '''
        cursor.execute(query, (id_usuario, asunto, observaciones, fecha_reporte))
        conn.commit()

        return jsonify({'mensaje': 'Reporte creado correctamente'}), 201

    except Exception as e:
        print(f'❌ ERROR EN /api/reportes: {e}')
        return jsonify({'error': 'Error al crear el reporte'}), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



@app.route('/api/asuntos', methods=['GET'])
def obtener_asuntos():
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT idAsunto, TipoAsunto FROM Asunto")
    resultados = cursor.fetchall()

    asuntos = [{'id': row[0], 'texto': row[1]} for row in resultados]

    cursor.close()
    conn.close()

    return jsonify(asuntos)


@app.route('/api/actas', methods=['POST'])
def crear_acta():
    try:
        data = request.json
        id_usuario = data.get('idUsuario')
        id_asunto = data.get('idAsunto')
        fecha_acta = data.get('FechaActa')
        comentario = data.get('Comentario')

        # Validación básica
        if not id_usuario or not id_asunto or not fecha_acta:
            return jsonify({"error": "Faltan datos obligatorios"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO ActaAdministrativa (idUsuario, idAsunto, FechaActa, Comentario)
            VALUES (?, ?, ?, ?)
        """, (id_usuario, id_asunto, fecha_acta, comentario))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({"message": "Acta administrativa creada correctamente"}), 201

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/solicitudes-aprobadas-rechazadas', methods=['GET'])
def obtener_solicitudes_aprobadas_rechazadas():
    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Consulta de vacaciones con nombre de usuario
        query_vacaciones = '''
            SELECT 
                v.idVacaciones, v.Usuario_idUsuario,
                v.EstadoSolicitud_idSolicitud, v.DiasSolicitados,
                v.FechaSalida, v.FechaRegreso,
                u.Nombres, u.Paterno, u.Materno
            FROM Vacaciones v
            JOIN Usuario u ON v.Usuario_idUsuario = u.idUsuario
            WHERE v.Estado is null
        '''
        cursor.execute(query_vacaciones)
        vacaciones = [
            {
                'tipo': 'Vacaciones',
                'id': row.idVacaciones,
                'usuario_id': row.Usuario_idUsuario,
                'nombre': f"{row.Nombres} {row.Paterno} {row.Materno}",
                'estado_id': row.EstadoSolicitud_idSolicitud,
                'dias_solicitados': row.DiasSolicitados,
                'fecha_salida': str(row.FechaSalida),
                'fecha_regreso': str(row.FechaRegreso)
            }
            for row in cursor.fetchall()
        ]

        # Consulta de permisos con nombre de usuario
        query_permisos = '''
            SELECT 
                p.idPermiso, p.Usuario_idUsuario,
                p.EstadoSolicitud_idSolicitud,
                p.DiaSolicitado, p.HoraInicio, p.HoraFin,
                p.Razon, c.TipoCompensacion,
                u.Nombres, u.Paterno, u.Materno
            FROM Permiso p
            LEFT JOIN Compensacion c ON p.Compensacion_idCompensacion = c.idCompensacion
            JOIN Usuario u ON p.Usuario_idUsuario = u.idUsuario
            WHERE p.Estado is null
        '''
        cursor.execute(query_permisos)
        permisos = [
    {
        'tipo': 'Permiso',
        'id': row.idPermiso,
        'usuario_id': row.Usuario_idUsuario,
        'estado_id': row.EstadoSolicitud_idSolicitud,
        'dia_solicitado': str(row.DiaSolicitado),
        'hora_inicio': str(row.HoraInicio),
        'hora_fin': str(row.HoraFin),
        'razon': row.Razon,
        'tipo_compensacion': row.TipoCompensacion,
        'nombre': f"{row.Nombres} {row.Paterno} {row.Materno}"
    }
    for row in cursor.fetchall()
]


        solicitudes = vacaciones + permisos
        return jsonify(solicitudes)

    except Exception as e:
        print("❌ Error en /solicitudes-aprobadas-rechazadas:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if conn:
            conn.close()

@app.route('/api/dar-baja-solicitud', methods=['POST'])
def dar_de_baja_solicitud():
    data = request.get_json()
    tipo = data.get("tipo")
    id_solicitud = data.get("id")

    if tipo not in ("Vacaciones", "Permiso") or not id_solicitud:
        return jsonify({"error": "Datos inválidos"}), 400

    conn = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Ejecutar update directo según tipo
        if tipo == "Permiso":
            cursor.execute("""
                UPDATE Permiso
                SET estado = 'Enterado'
                WHERE idPermiso = ?
            """, (id_solicitud,))

        elif tipo == "Vacaciones":
            cursor.execute("""
                UPDATE Vacaciones
                SET estado = 'Enterado'
                WHERE idVacaciones = ?
            """, (id_solicitud,))

        conn.commit()
        return jsonify({"mensaje": "Solicitud dada de baja correctamente"})

    except Exception as e:
        print("❌ Error al dar de baja solicitud:", e)
        return jsonify({"error": str(e)}), 500

    finally:
        if conn:
            conn.close()



@app.route('/api/usuarios/buscar', methods=['GET'])
def buscar_usuarios():
    texto = request.args.get("q", "")

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT idUsuario, nombres, paterno, materno
            FROM Usuario
            WHERE CONCAT(nombres, ' ', paterno, ' ', materno) LIKE ?
        """, f'%{texto}%')
        rows = cursor.fetchall()

        usuarios = [
            {
                "idUsuario": row[0],
                "nombreCompleto": f"{row[1]} {row[2]} {row[3]}"
            }
            for row in rows
        ]

        return jsonify(usuarios)

    except Exception as e:
        print(f"Error buscando usuarios: {e}")
        return jsonify([]), 500
    
@app.route('/api/reportes-usuarios', methods=['GET'])
def obtener_reportes_usuarios():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                u.Nombres, 
                u.Paterno, 
                u.Materno, 
                a.TipoAsunto,        
                r.Observaciones, 
                r.FechaReporte,     
                r.idReporte
            FROM Usuario u
            INNER JOIN Reporte r ON u.idUsuario = r.Usuario_idUsuario
            INNER JOIN Asunto a ON r.Asunto_idAsunto = a.idAsunto
            WHERE r.Estado IS NULL;
        """)
        resultados = cursor.fetchall()

        # Construir lista de diccionarios para JSON
        lista_reportes = []
        for fila in resultados:
            lista_reportes.append({
                "Nombres": fila[0],
                "Paterno": fila[1],
                "Materno": fila[2],
                "Asunto": fila[3],
                "Observaciones": fila[4],
                "FechaReporte": fila[5].strftime('%Y-%m-%d') if fila[5] else None,  # Formatear fecha
                "idReporte": fila[6]
            })

        return jsonify(lista_reportes), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

@app.route('/api/actas', methods=['GET'])
def obtener_actas():
    try:
        id_usuario = request.args.get('idUsuario')
        if not id_usuario:
            return jsonify({"error": "Se requiere el idUsuario"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Obtener el rol del usuario
        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Usuario no encontrado"}), 404

        tipo_rol = rol_row[0]

        if tipo_rol in [2, 3]:  # Admin o RH - Ver todas
            query = """
                SELECT 
                    u.nombres, 
                    u.materno, 
                    u.paterno, 
                    a.TipoAsunto, 
                    act.FechaActa, 
                    act.Comentario
                FROM ActaAdministrativa act
                INNER JOIN Usuario u ON act.idUsuario = u.idUsuario
                INNER JOIN Asunto a ON act.idAsunto = a.idAsunto
            """
            cursor.execute(query)
            
        elif tipo_rol == 4:  # Líder - Ver solo de su área
            cursor.execute("""
                SELECT DISTINCT ua.idArea 
                FROM Usuario_Area ua 
                WHERE ua.idUsuario = ?
            """, (id_usuario,))
            areas = cursor.fetchall()
            
            if not areas:
                return jsonify([]), 200
            
            area_ids = [str(area[0]) for area in areas]
            placeholders = ",".join("?" for _ in area_ids)
            
            query = f"""
                SELECT DISTINCT
                    u.nombres, 
                    u.materno, 
                    u.paterno, 
                    a.TipoAsunto, 
                    act.FechaActa, 
                    act.Comentario
                FROM ActaAdministrativa act
                INNER JOIN Usuario u ON act.idUsuario = u.idUsuario
                INNER JOIN Asunto a ON act.idAsunto = a.idAsunto
                INNER JOIN Usuario_Area ua ON ua.idUsuario = u.idUsuario
                WHERE ua.idArea IN ({placeholders})
                  AND u.Rol_idRol = 1
            """
            cursor.execute(query, area_ids)
            
        else:  # Empleado - Ver solo sus propias actas
            query = """
                SELECT 
                    u.nombres, 
                    u.materno, 
                    u.paterno, 
                    a.TipoAsunto, 
                    act.FechaActa, 
                    act.Comentario
                FROM ActaAdministrativa act
                INNER JOIN Usuario u ON act.idUsuario = u.idUsuario
                INNER JOIN Asunto a ON act.idAsunto = a.idAsunto
                WHERE u.idUsuario = ?
            """
            cursor.execute(query, (id_usuario,))

        columns = [column[0] for column in cursor.description]
        actas = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return jsonify(actas), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



@app.route('/api/reportes/<int:id_reporte>/enterado', methods=['PUT'])
def marcar_reporte_enterado(id_reporte):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE Reporte
            SET Estado = 'Enterado'
            WHERE idReporte = ?
        """, id_reporte)
        conn.commit()
        return jsonify({'mensaje': 'Reporte marcado como enterado'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

    finally:
        cursor.close()
        conn.close()


@app.route('/api/totalSolicitudes/<int:idUsuario>')
def total_solicitudes_por_rol(idUsuario):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Obtener el rol del usuario
        cursor.execute("""
            SELECT r.TipoRol 
            FROM Usuario u 
            JOIN Rol r ON u.Rol_idRol = r.idRol 
            WHERE u.idUsuario = ?
        """, (idUsuario,))
        resultado = cursor.fetchone()

        if not resultado:
            return jsonify({"error": "Usuario no encontrado"}), 404

        tipo_rol = resultado[0].lower()

        if tipo_rol == 'administrador':
            # El admin ve solicitudes en estados 18, 19 y 20
            cursor.execute("""
                SELECT 
                    (SELECT COUNT(*) FROM Permiso WHERE EstadoSolicitud_idSolicitud IN (18,19,20)) +
                    (SELECT COUNT(*) FROM Vacaciones WHERE EstadoSolicitud_idSolicitud IN (18,19,20))
            """)
            total = cursor.fetchone()[0]
            return jsonify({"total": total})

        elif tipo_rol == 'rh':
            estado = 'Pendiente de aprobar por Recursos Humanos'
        elif tipo_rol == 'lider area':
            estado = 'Pendiente de aprobar por tu líder'
        else:
            return jsonify({"total": 0})

        # Obtener el ID del estado correspondiente
        cursor.execute("""
            SELECT idSolicitud FROM EstadoSolicitud WHERE Estado = ?
        """, (estado,))
        estado_result = cursor.fetchone()

        if not estado_result:
            return jsonify({"total": 0})

        id_estado = estado_result[0]

        if tipo_rol == 'rh':
            # RH ve todas las solicitudes con ese estado
            cursor.execute("""
                SELECT 
                    (SELECT COUNT(*) FROM Permiso WHERE EstadoSolicitud_idSolicitud = ?) +
                    (SELECT COUNT(*) FROM Vacaciones WHERE EstadoSolicitud_idSolicitud = ?)
            """, (id_estado, id_estado))
            total = cursor.fetchone()[0]
            return jsonify({"total": total})

        elif tipo_rol == 'lider area':
            # Obtener áreas del líder
            cursor.execute("""
                SELECT idArea 
                FROM Usuario_Area 
                WHERE idUsuario = ?
            """, (idUsuario,))
            areas = [row[0] for row in cursor.fetchall()]

            if not areas:
                return jsonify({"total": 0})

            placeholders = ','.join(['?'] * len(areas))

            # Contar permisos y vacaciones solo de usuarios en sus áreas y con estado correspondiente
            query = f"""
                 SELECT 
    (SELECT COUNT(DISTINCT p.idPermiso) FROM Permiso p
     JOIN Usuario u ON p.Usuario_idUsuario = u.idUsuario
     JOIN Usuario_Area ua ON u.idUsuario = ua.idUsuario
     WHERE ua.idArea IN ({placeholders}) AND p.EstadoSolicitud_idSolicitud = ?) +
    (SELECT COUNT(DISTINCT v.idVacaciones) FROM Vacaciones v
     JOIN Usuario u ON v.Usuario_idUsuario = u.idUsuario
     JOIN Usuario_Area ua ON u.idUsuario = ua.idUsuario
     WHERE ua.idArea IN ({placeholders}) AND v.EstadoSolicitud_idSolicitud = ?)


            """
            params = areas + [id_estado] + areas + [id_estado]
            cursor.execute(query, params)
            total = cursor.fetchone()[0]

            return jsonify({"total": total})

    except Exception as e:
        print(f"Error en total_solicitudes_por_rol: {str(e)}")
        return jsonify({"error": "Error interno del servidor"}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/reportesPendientes', methods=['GET'])
def obtener_reportes_pendientes():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM Reporte WHERE Estado IS NULL")
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        # Devolvemos el conteo en un objeto JSON
        return jsonify({"pendientes": row[0]})

    except Exception as e:
        print(f"Error al obtener reportes pendientes: {e}")
        return jsonify({"pendientes": 0}), 500




@app.route('/vacaciones_por_ley', methods=['POST'])
def vacaciones_por_ley():
    hoy = datetime.now().date()
    conn = get_connection()
    cursor = conn.cursor()

    # Obtener todos los usuarios con fecha de ingreso, vacaciones actuales y días disponibles
    cursor.execute("SELECT idUsuario, FechaIngreso, Vacaciones, DiasDisponibles FROM Usuario")
    usuarios = cursor.fetchall()

    actualizados = []

    for idUsuario, fechaIngreso, vacacionesActuales, diasDisponibles in usuarios:
        if not fechaIngreso:
            continue

        # Calcular los días que deberían tener por ley
        vacacionesCalculadas = calcular_dias_vacaciones(fechaIngreso.strftime('%Y-%m-%d'))

        # Verificar si hay cambio en las vacaciones por ley
        if vacacionesActuales != vacacionesCalculadas:
            # Calcular nuevos días disponibles según las reglas
            if diasDisponibles < 0:
                # Si debe días, restamos los días negativos de las nuevas vacaciones
                nuevos_dias_disponibles = vacacionesCalculadas + diasDisponibles
            else:
                # Si tenía días positivos, no se acumulan, se pierden
                nuevos_dias_disponibles = vacacionesCalculadas
            
            # Asegurarse que no queden días negativos (por si debe más días de los nuevos disponibles)
            nuevos_dias_disponibles = max(nuevos_dias_disponibles, 0)

            # Actualizar ambos campos
            cursor.execute(
                "UPDATE Usuario SET Vacaciones = ?, DiasDisponibles = ? WHERE idUsuario = ?",
                (vacacionesCalculadas, nuevos_dias_disponibles, idUsuario)
            )
            
            actualizados.append({
                'idUsuario': idUsuario,
                'nuevas_vacaciones': vacacionesCalculadas,
                'nuevos_dias_disponibles': nuevos_dias_disponibles,
                'dias_disponibles_anteriores': diasDisponibles
            })

    conn.commit()


    conn.close()

    return jsonify({
        'mensaje': f'{len(actualizados)} usuario(s) actualizado(s)',
        'usuarios_actualizados': actualizados
    })



@app.route('/api/tiposIncapacidad', methods=['GET'])
def obtener_tipos_incapacidad():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT idTipoIncapacidad, TipoIncapacidad FROM TipoIncapacidad")
        rows = cursor.fetchall()

        # Cambia "id" por "idTipoIncapacidad" para coincidir con el frontend
        tipos = [
            {
                "idTipoIncapacidad": row[0],  # Cambiado de "id"
                "nombre": row[1]
            }
            for row in rows
        ]

        return jsonify(tipos)

    except Exception as e:
        print(f"Error al obtener tipos de incapacidad: {e}")
        return jsonify([]), 500


@app.route('/api/registrarIncapacidad', methods=['POST'])
def registrar_incapacidad():
    data = request.get_json()
    print("Datos recibidos:", data)  # Para depuración

    # Validación mejorada
    required_fields = {
        'idTipoIncapacidad': int,
        'idUsuario': int,
        'fechaInicio': str,
        'fechaFinal': str,
        'observaciones': str
    }
    
    errors = []
    cleaned_data = {}
    
    for field, field_type in required_fields.items():
        value = data.get(field)
        if value is None:
            errors.append(f"Campo '{field}' es requerido")
        else:
            try:
                cleaned_data[field] = field_type(value)
            except (ValueError, TypeError):
                errors.append(f"Campo '{field}' debe ser {field_type.__name__}")
    
    if errors:
        return jsonify({"error": "Validación fallida", "details": errors}), 400

    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Query principal de inserción
        insert_query = """
            INSERT INTO Incapacidad (
                TipoIncapacidad_idTipoIncapacidad,
                Usuario_idUsuario,
                fechaInicio,
                fechaFinal,
                Observaciones
            ) VALUES (?, ?, ?, ?, ?)
        """
        
        # Parámetros para el INSERT
        params = (
            cleaned_data['idTipoIncapacidad'],
            cleaned_data['idUsuario'],
            cleaned_data['fechaInicio'],
            cleaned_data['fechaFinal'],
            cleaned_data['observaciones']
        )
        
        # Ejecutar inserción
        cursor.execute(insert_query, params)
        
        # Obtener el ID insertado (forma correcta para SQL Server)
        cursor.execute("SELECT SCOPE_IDENTITY() AS new_id")
        new_id = cursor.fetchone()[0]
        
        conn.commit()

        return jsonify({
            "message": "Incapacidad registrada correctamente",
            "id": new_id  # Retorna el ID generado
        }), 200

    except Exception as e:
        print("Error al registrar incapacidad:", str(e))
        if conn:
            conn.rollback()
        return jsonify({
            "error": "Error interno al registrar la incapacidad",
            "details": str(e)
        }), 500
    finally:
        # Cerrar cursor y conexión de forma segura
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/incapacidades', methods=['GET'])
def obtener_incapacidades():
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        query = """
            SELECT 
                i.idIncapacidad,           
                u.idUsuario,               
                ti.tipoIncapacidad AS TipoDeIncapacidad,
                u.Nombres,
                u.Paterno,
                u.Materno,
                i.fechaInicio,
                i.fechaFinal,
                i.Observaciones
            FROM Incapacidad i
            JOIN Usuario u ON u.idUsuario = i.Usuario_idUsuario
            JOIN TipoIncapacidad ti ON ti.idTipoIncapacidad = i.TipoIncapacidad_idTipoIncapacidad
            WHERE i.Estado IS NULL
        """

        cursor.execute(query)
        rows = cursor.fetchall()

        incapacidades = []
        for row in rows:
            incapacidades.append({
                "idIncapacidad": row.idIncapacidad,    
                "idUsuario": row.idUsuario,             
                "tipo": row.TipoDeIncapacidad,
                "nombres": row.Nombres,
                "paterno": row.Paterno,
                "materno": row.Materno,
                "fechaInicio": row.fechaInicio.strftime('%Y-%m-%d') if row.fechaInicio else None,
                "fechaFinal": row.fechaFinal.strftime('%Y-%m-%d') if row.fechaFinal else None,
                "observaciones": row.Observaciones
            })

        return jsonify(incapacidades), 200

    except Exception as e:
        print("Error al obtener incapacidades:", str(e))
        return jsonify({
            "error": "Error interno al obtener las incapacidades",
            "details": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/incapacidades/<int:idIncapacidad>/enterado', methods=['PUT'])
def marcar_incapacidad_enterado(idIncapacidad):
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Actualizar por ID específico de incapacidad
        query = """
            UPDATE Incapacidad
            SET Estado = 'Enterado'
            WHERE idIncapacidad = ?
              AND Estado IS NULL
        """
        cursor.execute(query, (idIncapacidad,))
        conn.commit()

        if cursor.rowcount > 0:
            return jsonify({"mensaje": "Incapacidad marcada como enterada"}), 200
        else:
            return jsonify({"mensaje": "No se encontró incapacidad pendiente con ese ID"}), 404

    except Exception as e:
        print("Error al actualizar incapacidad:", str(e))
        return jsonify({
            "error": "Error interno al actualizar la incapacidad",
            "details": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.route('/api/usuario/<int:idUsuario>/rol', methods=['GET'])
def obtener_rol_usuario(idUsuario):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", idUsuario)
        row = cursor.fetchone()

        if row:
            return jsonify({"idRol": row[0]})
        else:
            return jsonify({"error": "Usuario no encontrado"}), 404

    except Exception as e:
        print("Error al obtener rol:", e)
        return jsonify({"error": "Error del servidor"}), 500

    finally:
        if conn:
            conn.close()

@app.route('/api/usuario/<int:idUsuario>/estado', methods=['GET'])
def obtener_estado_usuario(idUsuario):
    try:
        conn = get_connection()
        cursor = conn.cursor()

        query = """
            SELECT 
                CASE 
                    WHEN Estado = 'Activo' THEN CAST(1 AS BIT)
                    WHEN Estado = 'Inactivo' THEN CAST(0 AS BIT)
                END AS EsActivo,
                FechaBaja,
                ComentarioSalida,
                CASE 
                    WHEN FechaBaja IS NOT NULL AND ComentarioSalida IS NOT NULL THEN CAST(1 AS BIT)
                    ELSE CAST(0 AS BIT)
                END AS EsBaja
            FROM Usuario
            WHERE idUsuario = ?
        """
        cursor.execute(query, idUsuario)
        row = cursor.fetchone()

        if row:
            return jsonify({
                "esActivo": int(row[0]),
                "fechaBaja": row[1],
                "comentarioSalida": row[2],
                "esBaja": int(row[3])
            })
        else:
            return jsonify({"error": "Usuario no encontrado"}), 404

    except Exception as e:
        print("Error al obtener estado:", e)
        return jsonify({"error": "Error del servidor"}), 500

    finally:
        if conn:
            conn.close()




@app.route('/api/incapacidadesExcel', methods=['GET'])
def exportar_incapacidades_excel():
    conn = None
    cursor = None
    try:
        conn = get_connection()
        cursor = conn.cursor()

        query = """
            SELECT 
                ti.tipoIncapacidad AS TipoDeIncapacidad,
                u.Nombres,
                u.Paterno,
                u.Materno,
                i.fechaInicio,
                i.fechaFinal,
                i.Observaciones
            FROM Incapacidad i
            JOIN Usuario u ON u.idUsuario = i.Usuario_idUsuario
            JOIN TipoIncapacidad ti ON ti.idTipoIncapacidad = i.TipoIncapacidad_idTipoIncapacidad
        """

        cursor.execute(query)
        rows = cursor.fetchall()

        # Acceso por índices y concatenación de nombre completo
        data = [{
            "Tipo de Incapacidad": row[0],
            "Empleado": f"{row[1]} {row[2]} {row[3]}".strip(),
            "Fecha de Inicio": row[4].strftime('%Y-%m-%d') if row[4] else "",
            "Fecha Final": row[5].strftime('%Y-%m-%d') if row[5] else "",
            "Observaciones": row[6] or ""
        } for row in rows]

        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Incapacidades')

            workbook = writer.book
            hoja = writer.sheets['Incapacidades']

            # Estilo visual igual que en /exportar-reportes-vista
            font = Font(bold=True, size=12, color="000000")  # Letra negra
            fill = PatternFill(start_color="f5fc19", end_color="f5fc19", fill_type="solid")  # Amarillo claro

            for col_num, column_title in enumerate(df.columns, 1):
                celda = hoja[f"{get_column_letter(col_num)}1"]
                celda.font = font
                celda.fill = fill
                hoja.column_dimensions[get_column_letter(col_num)].width = max(15, len(column_title) + 5)

        output.seek(0)
        return send_file(
            output,
            download_name="Incapacidades.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("Error al exportar Excel:", str(e))
        return jsonify({
            "error": "Error al exportar Excel",
            "details": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/graficos/distribucion', methods=['GET'])
def graficos_distribucion():
    conn = get_connection()
    cursor = conn.cursor()

    # --- Distribución por Rol ---
    cursor.execute("""
        SELECT r.TipoRol, COUNT(*) as total
        FROM Usuario u
        JOIN Rol r ON u.Rol_idRol = r.idRol
        WHERE u.Estado = 'Activo'
        GROUP BY r.TipoRol
    """)
    roles = cursor.fetchall()
    data_roles = [{"label": r[0], "count": r[1]} for r in roles]

    # --- Distribución por Área ---
    cursor.execute("""
        SELECT a.NombreArea, COUNT(*) as total
        FROM Usuario_Area ua
        JOIN Area a ON ua.idArea = a.idArea
        JOIN Usuario u ON u.idUsuario = ua.idUsuario
        WHERE u.Estado = 'Activo'
        GROUP BY a.NombreArea
    """)
    areas = cursor.fetchall()
    data_areas = [{"label": a[0], "count": a[1]} for a in areas]

    # --- Reportes por Área (área principal) ---
    cursor.execute("""
        ;WITH AreaPrincipalPorUsuario AS (
            SELECT ua.idUsuario, MIN(ua.idArea) AS idAreaPrincipal
            FROM Usuario_Area ua
            GROUP BY ua.idUsuario
        )
        SELECT a.NombreArea, COUNT(r.idReporte) AS TotalReportes
        FROM Reporte r
        JOIN Usuario u ON r.Usuario_idUsuario = u.idUsuario
        JOIN AreaPrincipalPorUsuario apu ON u.idUsuario = apu.idUsuario
        JOIN Area a ON apu.idAreaPrincipal = a.idArea
        GROUP BY a.NombreArea
        ORDER BY TotalReportes DESC;
    """)
    reportes = cursor.fetchall()
    data_reportes = [{"label": r[0], "count": r[1]} for r in reportes]

    # --- Vacaciones por Área ---
    cursor.execute("""
        ;WITH AreaPrincipalPorUsuario AS (
            SELECT ua.idUsuario, MIN(ua.idArea) AS idAreaPrincipal
            FROM Usuario_Area ua
            GROUP BY ua.idUsuario
        )
        SELECT a.NombreArea, COUNT(v.idVacaciones) AS TotalVacaciones
        FROM Vacaciones v
        JOIN Usuario u ON v.Usuario_idUsuario = u.idUsuario
        JOIN AreaPrincipalPorUsuario apu ON u.idUsuario = apu.idUsuario
        JOIN Area a ON apu.idAreaPrincipal = a.idArea
        GROUP BY a.NombreArea
        ORDER BY TotalVacaciones DESC;
    """)
    vacaciones = cursor.fetchall()
    data_vacaciones = [{"label": v[0], "count": v[1]} for v in vacaciones]

    # --- Permisos por Área ---
    cursor.execute("""
        ;WITH AreaPrincipalPorUsuario AS (
            SELECT ua.idUsuario, MIN(ua.idArea) AS idAreaPrincipal
            FROM Usuario_Area ua
            GROUP BY ua.idUsuario
        )
        SELECT a.NombreArea, COUNT(p.idPermiso) AS TotalPermisos
        FROM Permiso p
        JOIN Usuario u ON p.Usuario_idUsuario = u.idUsuario
        JOIN AreaPrincipalPorUsuario apu ON u.idUsuario = apu.idUsuario
        JOIN Area a ON apu.idAreaPrincipal = a.idArea
        GROUP BY a.NombreArea
        ORDER BY TotalPermisos DESC;
    """)
    permisos = cursor.fetchall()
    data_permisos = [{"label": p[0], "count": p[1]} for p in permisos]

    # --- Ausentismo por Área (año actual) ---
    cursor.execute("""
        ;WITH AreaPrincipalPorUsuario AS (
            SELECT ua.idUsuario, MIN(ua.idArea) AS idAreaPrincipal
            FROM Usuario_Area ua
            GROUP BY ua.idUsuario
        )
        SELECT a.NombreArea, COUNT(r.idReporte) AS TotalAusentismo
        FROM Reporte r
        JOIN Asunto s ON r.Asunto_idAsunto = s.idAsunto
        JOIN Usuario u ON r.Usuario_idUsuario = u.idUsuario
        JOIN AreaPrincipalPorUsuario apu ON u.idUsuario = apu.idUsuario
        JOIN Area a ON apu.idAreaPrincipal = a.idArea
        WHERE s.TipoAsunto = 'Ausentismo' AND YEAR(r.FechaReporte) = YEAR(GETDATE())
        GROUP BY a.NombreArea
        ORDER BY TotalAusentismo DESC;
    """)
    ausentismo = cursor.fetchall()
    data_ausentismo = [{"label": r[0], "count": r[1]} for r in ausentismo]

    # --- Retardos por Área (año actual) ---
    cursor.execute("""
        ;WITH AreaPrincipalPorUsuario AS (
            SELECT ua.idUsuario, MIN(ua.idArea) AS idAreaPrincipal
            FROM Usuario_Area ua
            GROUP BY ua.idUsuario
        )
        SELECT a.NombreArea, COUNT(r.idReporte) AS TotalRetardos
        FROM Reporte r
        JOIN Asunto s ON r.Asunto_idAsunto = s.idAsunto
        JOIN Usuario u ON r.Usuario_idUsuario = u.idUsuario
        JOIN AreaPrincipalPorUsuario apu ON u.idUsuario = apu.idUsuario
        JOIN Area a ON apu.idAreaPrincipal = a.idArea
        WHERE s.TipoAsunto = 'Retardo' AND YEAR(r.FechaReporte) = YEAR(GETDATE())
        GROUP BY a.NombreArea
        ORDER BY TotalRetardos DESC;
    """)
    retardos = cursor.fetchall()
    data_retardos = [{"label": r[0], "count": r[1]} for r in retardos]

    # --- Distribución por Razón de Baja ---
    cursor.execute("""
        SELECT rb.RazonBaja, COUNT(u.idUsuario) AS TotalUsuarios
        FROM RazonesBaja rb
        INNER JOIN Usuario u ON rb.idRazonBaja = u.idRazonBaja
        GROUP BY rb.RazonBaja
    """)
    razones_baja = cursor.fetchall()
    data_razones_baja = [{"label": r[0], "count": r[1]} for r in razones_baja]

    # --- Retornar todo junto ---
    return jsonify({
        "roles": data_roles,
        "areas": data_areas,
        "reportes_por_area": data_reportes,
        "vacaciones_por_area": data_vacaciones,
        "permisos_por_area": data_permisos,
        "ausentismo_por_area": data_ausentismo,
        "retardos_por_area": data_retardos,
        "razones_baja": data_razones_baja
    })


@app.route('/graficos')
def vista_graficos():
    return render_template('stats.html')

@app.route('/api/empleados/mes/<int:mes>', methods=['GET'])
def empleados_por_mes(mes):
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT idUsuario, Nombres, Paterno, Materno, FechaIngreso
            FROM Usuario
            WHERE MONTH(FechaIngreso) = ?
            ORDER BY FechaIngreso DESC
        """, mes)
        
        rows = cursor.fetchall()
        columnas = [col[0] for col in cursor.description]
        empleados = [dict(zip(columnas, row)) for row in rows]

        from datetime import datetime
        for emp in empleados:
            fecha_ingreso = emp.get('FechaIngreso')
            if isinstance(fecha_ingreso, datetime):
                # Convertir a string YYYY-MM-DD para evitar problemas de zona horaria
                emp['FechaIngreso'] = fecha_ingreso.strftime('%Y-%m-%d')
            else:
                # Asegura formato consistente si viene como string
                emp['FechaIngreso'] = str(fecha_ingreso)[:10]

        return jsonify({'empleados': empleados}), 200

    except Exception as e:
        print(f'❌ ERROR en /api/empleados/mes/{mes}: {e}')
        return jsonify({'error': 'Error al obtener los empleados'}), 500

    finally:
        if cursor: cursor.close()
        if conn: conn.close()

import signal
from contextlib import contextmanager

# Función de timeout
class TimeoutError(Exception):
    pass

def timeout_handler(signum, frame):
    raise TimeoutError("Query timeout")

@contextmanager
def query_timeout(seconds):
    # Configurar el timeout
    old_handler = signal.signal(signal.SIGALRM, timeout_handler)
    signal.alarm(seconds)
    try:
        yield
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, old_handler)

@app.route('/api/reportes')
def get_reportes():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT r.idReporte, u.Nombres, u.Paterno, u.Materno,
                   r.Observaciones, a.TipoAsunto, r.FechaReporte
            FROM Reporte r
            INNER JOIN Usuario u ON u.idUsuario = r.Usuario_idUsuario
            INNER JOIN Asunto a ON a.idAsunto = r.Asunto_idAsunto
        """)
        
        reportes = []
        for row in cursor.fetchall():
            reportes.append({
                'id': row[0],
                'empleado': f"{row[1]} {row[2]} {row[3]}",
                'observaciones': row[4] or '',
                'tipo_asunto': row[5] or '',
                'fecha': row[6].isoformat() if row[6] else ''
            })
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'data': reportes})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/vacaciones')
def get_vacaciones():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT v.idVacaciones, u.Nombres, u.Paterno, u.Materno,
                   e.Estado AS EstadoSolicitud, v.DiasSolicitados,
                   v.FechaSalida, v.FechaRegreso
            FROM Vacaciones v
            INNER JOIN Usuario u ON u.idUsuario = v.Usuario_idUsuario
            INNER JOIN EstadoSolicitud e ON e.idSolicitud = v.EstadoSolicitud_idSolicitud
        """)
        
        vacaciones = []
        for row in cursor.fetchall():
            vacaciones.append({
                'id': row[0],
                'empleado': f"{row[1]} {row[2]} {row[3]}",
                'estado': row[4] or '',
                'dias_solicitados': row[5] or 0,
                'fecha_salida': row[6].isoformat() if row[6] else '',
                'fecha_regreso': row[7].isoformat() if row[7] else ''
            })
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'data': vacaciones})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/permisos')
def get_permisos():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT p.idPermiso, u.Nombres, u.Paterno, u.Materno,
                   e.Estado AS EstadoSolicitud, p.DiaSolicitado, p.HoraInicio, p.HoraFin,
                   p.Razon, c.TipoCompensacion
            FROM Permiso p
            INNER JOIN Usuario u ON u.idUsuario = p.Usuario_idUsuario
            INNER JOIN Compensacion c ON c.idCompensacion = p.Compensacion_idCompensacion
            INNER JOIN EstadoSolicitud e ON e.idSolicitud = p.EstadoSolicitud_idSolicitud
        """)
        
        permisos = []
        for row in cursor.fetchall():
            permisos.append({
                'id': row[0],
                'empleado': f"{row[1]} {row[2]} {row[3]}",
                'estado': row[4] or '',
                'dia_solicitado': row[5].isoformat() if row[5] else '',
                'hora_inicio': row[6].strftime('%H:%M') if row[6] else '',
                'hora_fin': row[7].strftime('%H:%M') if row[7] else '',
                'razon': row[8] or '',
                'tipo_compensacion': row[9] or ''
            })
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'data': permisos})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/reportes-historial')
def get_reportes_historial():
    try:
        id_usuario = request.args.get('idUsuario')
        if not id_usuario:
            return jsonify({"error": "Se requiere el idUsuario"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Obtener el rol del usuario
        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Usuario no encontrado"}), 404

        tipo_rol = rol_row[0]

        if tipo_rol in [2, 3]:  # Admin o RH - Ver todos
            query = """
                SELECT r.idReporte, u.Nombres, u.Paterno, u.Materno,
                       r.Observaciones, a.TipoAsunto, r.FechaReporte
                FROM Reporte r
                INNER JOIN Usuario u ON u.idUsuario = r.Usuario_idUsuario
                INNER JOIN Asunto a ON a.idAsunto = r.Asunto_idAsunto
            """
            cursor.execute(query)
            
        elif tipo_rol == 4:  # Líder - Ver solo de su área
            # Obtener áreas del líder
            cursor.execute("""
                SELECT DISTINCT ua.idArea 
                FROM Usuario_Area ua 
                WHERE ua.idUsuario = ?
            """, (id_usuario,))
            areas = cursor.fetchall()
            
            if not areas:
                return jsonify([]), 200
            
            area_ids = [str(area[0]) for area in areas]
            placeholders = ",".join("?" for _ in area_ids)
            
            query = f"""
                SELECT DISTINCT r.idReporte, u.Nombres, u.Paterno, u.Materno,
                       r.Observaciones, a.TipoAsunto, r.FechaReporte
                FROM Reporte r
                INNER JOIN Usuario u ON u.idUsuario = r.Usuario_idUsuario
                INNER JOIN Asunto a ON a.idAsunto = r.Asunto_idAsunto
                INNER JOIN Usuario_Area ua ON ua.idUsuario = u.idUsuario
                WHERE ua.idArea IN ({placeholders})
                  AND u.Rol_idRol = 1
            """
            cursor.execute(query, area_ids)
            
        else:  # Empleado - Ver solo sus propios reportes
            query = """
                SELECT r.idReporte, u.Nombres, u.Paterno, u.Materno,
                       r.Observaciones, a.TipoAsunto, r.FechaReporte
                FROM Reporte r
                INNER JOIN Usuario u ON u.idUsuario = r.Usuario_idUsuario
                INNER JOIN Asunto a ON a.idAsunto = r.Asunto_idAsunto
                WHERE u.idUsuario = ?
            """
            cursor.execute(query, (id_usuario,))

        columns = [column[0] for column in cursor.description]
        reportes = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        return jsonify(reportes), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/vacaciones-historial')
def get_vacaciones_historial():
    try:
        id_usuario = request.args.get('idUsuario')
        if not id_usuario:
            return jsonify({"error": "Se requiere el idUsuario"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Obtener el rol del usuario
        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Usuario no encontrado"}), 404

        tipo_rol = rol_row[0]

        if tipo_rol in [2, 3]:  # Admin o RH - Ver todos
            query = """
                SELECT v.idVacaciones, u.Nombres, u.Paterno, u.Materno,
                       e.Estado AS EstadoSolicitud, v.DiasSolicitados,
                       v.FechaSalida, v.FechaRegreso
                FROM Vacaciones v
                INNER JOIN Usuario u ON u.idUsuario = v.Usuario_idUsuario
                INNER JOIN EstadoSolicitud e ON e.idSolicitud = v.EstadoSolicitud_idSolicitud
            """
            cursor.execute(query)
            
        elif tipo_rol == 4:  # Líder - Ver solo de su área
            cursor.execute("""
                SELECT DISTINCT ua.idArea 
                FROM Usuario_Area ua 
                WHERE ua.idUsuario = ?
            """, (id_usuario,))
            areas = cursor.fetchall()
            
            if not areas:
                return jsonify([]), 200
            
            area_ids = [str(area[0]) for area in areas]
            placeholders = ",".join("?" for _ in area_ids)
            
            query = f"""
                SELECT DISTINCT v.idVacaciones, u.Nombres, u.Paterno, u.Materno,
                       e.Estado AS EstadoSolicitud, v.DiasSolicitados,
                       v.FechaSalida, v.FechaRegreso
                FROM Vacaciones v
                INNER JOIN Usuario u ON u.idUsuario = v.Usuario_idUsuario
                INNER JOIN EstadoSolicitud e ON e.idSolicitud = v.EstadoSolicitud_idSolicitud
                INNER JOIN Usuario_Area ua ON ua.idUsuario = u.idUsuario
                WHERE ua.idArea IN ({placeholders})
                  AND u.Rol_idRol = 1
            """
            cursor.execute(query, area_ids)
            
        else:  # Empleado - Ver solo sus propias vacaciones
            query = """
                SELECT v.idVacaciones, u.Nombres, u.Paterno, u.Materno,
                       e.Estado AS EstadoSolicitud, v.DiasSolicitados,
                       v.FechaSalida, v.FechaRegreso
                FROM Vacaciones v
                INNER JOIN Usuario u ON u.idUsuario = v.Usuario_idUsuario
                INNER JOIN EstadoSolicitud e ON e.idSolicitud = v.EstadoSolicitud_idSolicitud
                WHERE u.idUsuario = ?
            """
            cursor.execute(query, (id_usuario,))

        columns = [column[0] for column in cursor.description]
        vacaciones = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        return jsonify(vacaciones), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/permisos-historial')
def get_permisos_historial():
    try:
        id_usuario = request.args.get('idUsuario')
        if not id_usuario:
            return jsonify({"error": "Se requiere el idUsuario"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Obtener el rol del usuario
        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Usuario no encontrado"}), 404

        tipo_rol = rol_row[0]

        if tipo_rol in [2, 3]:  # Admin o RH - Ver todos
            query = """
                SELECT p.idPermiso, u.Nombres, u.Paterno, u.Materno,
                       e.Estado AS EstadoSolicitud, p.DiaSolicitado, p.HoraInicio, p.HoraFin,
                       p.Razon, c.TipoCompensacion
                FROM Permiso p
                INNER JOIN Usuario u ON u.idUsuario = p.Usuario_idUsuario
                INNER JOIN Compensacion c ON c.idCompensacion = p.Compensacion_idCompensacion
                INNER JOIN EstadoSolicitud e ON e.idSolicitud = p.EstadoSolicitud_idSolicitud
            """
            cursor.execute(query)
            
        elif tipo_rol == 4:  # Líder - Ver solo de su área
            cursor.execute("""
                SELECT DISTINCT ua.idArea 
                FROM Usuario_Area ua 
                WHERE ua.idUsuario = ?
            """, (id_usuario,))
            areas = cursor.fetchall()
            
            if not areas:
                return jsonify([]), 200
            
            area_ids = [str(area[0]) for area in areas]
            placeholders = ",".join("?" for _ in area_ids)
            
            query = f"""
                SELECT DISTINCT p.idPermiso, u.Nombres, u.Paterno, u.Materno,
                       e.Estado AS EstadoSolicitud, p.DiaSolicitado, p.HoraInicio, p.HoraFin,
                       p.Razon, ISNULL(c.TipoCompensacion, 'Sin compensación') AS TipoCompensacion
                FROM Permiso p
                INNER JOIN Usuario u ON u.idUsuario = p.Usuario_idUsuario
                LEFT JOIN Compensacion c ON c.idCompensacion = p.Compensacion_idCompensacion
                INNER JOIN EstadoSolicitud e ON e.idSolicitud = p.EstadoSolicitud_idSolicitud
                INNER JOIN Usuario_Area ua ON ua.idUsuario = u.idUsuario
                WHERE ua.idArea IN ({placeholders})
                  AND u.Rol_idRol = 1
            """
            cursor.execute(query, area_ids)
            
        else:  # Empleado - Ver solo sus propios permisos
            query = """
                SELECT p.idPermiso, u.Nombres, u.Paterno, u.Materno,
                       e.Estado AS EstadoSolicitud, p.DiaSolicitado, p.HoraInicio, p.HoraFin,
                       p.Razon, c.TipoCompensacion
                FROM Permiso p
                INNER JOIN Usuario u ON u.idUsuario = p.Usuario_idUsuario
                INNER JOIN Compensacion c ON c.idCompensacion = p.Compensacion_idCompensacion
                INNER JOIN EstadoSolicitud e ON e.idSolicitud = p.EstadoSolicitud_idSolicitud
                WHERE u.idUsuario = ?
            """
            cursor.execute(query, (id_usuario,))

        columns = [column[0] for column in cursor.description]
        permisos = []
        
        for row in cursor.fetchall():
            row_dict = {}
            for col_name, value in zip(columns, row):
                if isinstance(value, (date, datetime)):
                    row_dict[col_name] = value.isoformat()
                elif isinstance(value, time):
                    row_dict[col_name] = value.strftime("%H:%M:%S")
                elif isinstance(value, Decimal):
                    row_dict[col_name] = float(value)
                else:
                    row_dict[col_name] = value
            permisos.append(row_dict)

        return jsonify(permisos), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/incapacidades-historial', methods=['GET'])
def obtener_incapacidades_historial():
    try:
        id_usuario = request.args.get('idUsuario')
        if not id_usuario:
            return jsonify({"error": "Se requiere el idUsuario"}), 400

        conn = get_connection()
        cursor = conn.cursor()

        # Obtener el rol del usuario
        cursor.execute("SELECT Rol_idRol FROM Usuario WHERE idUsuario = ?", (id_usuario,))
        rol_row = cursor.fetchone()
        if not rol_row:
            return jsonify({"error": "Usuario no encontrado"}), 404

        tipo_rol = rol_row[0]

        if tipo_rol in [2, 3]:  # Admin o RH - Ver todas
            query = """
                SELECT 
                    i.idIncapacidad,           
                    u.idUsuario,               
                    ti.tipoIncapacidad AS TipoDeIncapacidad,
                    u.Nombres,
                    u.Paterno,
                    u.Materno,
                    i.fechaInicio,
                    i.fechaFinal,
                    i.Observaciones
                FROM Incapacidad i
                JOIN Usuario u ON u.idUsuario = i.Usuario_idUsuario
                JOIN TipoIncapacidad ti ON ti.idTipoIncapacidad = i.TipoIncapacidad_idTipoIncapacidad
            """
            cursor.execute(query)
            
        elif tipo_rol == 4:  # Líder - Ver solo de su área
            cursor.execute("""
                SELECT DISTINCT ua.idArea 
                FROM Usuario_Area ua 
                WHERE ua.idUsuario = ?
            """, (id_usuario,))
            areas = cursor.fetchall()
            
            if not areas:
                return jsonify([]), 200
            
            area_ids = [str(area[0]) for area in areas]
            placeholders = ",".join("?" for _ in area_ids)
            
            query = f"""
                SELECT DISTINCT
                    i.idIncapacidad,           
                    u.idUsuario,               
                    ti.tipoIncapacidad AS TipoDeIncapacidad,
                    u.Nombres,
                    u.Paterno,
                    u.Materno,
                    i.fechaInicio,
                    i.fechaFinal,
                    i.Observaciones
                FROM Incapacidad i
                JOIN Usuario u ON u.idUsuario = i.Usuario_idUsuario
                JOIN TipoIncapacidad ti ON ti.idTipoIncapacidad = i.TipoIncapacidad_idTipoIncapacidad
                INNER JOIN Usuario_Area ua ON ua.idUsuario = u.idUsuario
                WHERE ua.idArea IN ({placeholders})
                  AND u.Rol_idRol = 1
            """
            cursor.execute(query, area_ids)
            
        else:  # Empleado - Ver solo sus propias incapacidades
            query = """
                SELECT 
                    i.idIncapacidad,           
                    u.idUsuario,               
                    ti.tipoIncapacidad AS TipoDeIncapacidad,
                    u.Nombres,
                    u.Paterno,
                    u.Materno,
                    i.fechaInicio,
                    i.fechaFinal,
                    i.Observaciones
                FROM Incapacidad i
                JOIN Usuario u ON u.idUsuario = i.Usuario_idUsuario
                JOIN TipoIncapacidad ti ON ti.idTipoIncapacidad = i.TipoIncapacidad_idTipoIncapacidad
                WHERE u.idUsuario = ?
            """
            cursor.execute(query, (id_usuario,))

        rows = cursor.fetchall()

        incapacidades = []
        for row in rows:
            incapacidades.append({
                "idIncapacidad": row.idIncapacidad,    
                "idUsuario": row.idUsuario,             
                "tipo": row.TipoDeIncapacidad,
                "nombres": row.Nombres,
                "paterno": row.Paterno,
                "materno": row.Materno,
                "fechaInicio": row.fechaInicio.strftime('%Y-%m-%d') if row.fechaInicio else None,
                "fechaFinal": row.fechaFinal.strftime('%Y-%m-%d') if row.fechaFinal else None,
                "observaciones": row.Observaciones
            })

        return jsonify(incapacidades), 200

    except Exception as e:
        print("Error al obtener incapacidades:", str(e))
        return jsonify({
            "error": "Error interno al obtener las incapacidades",
            "details": str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@app.route('/api/exportar-reportes-vista', methods=['POST'])
def exportar_reportes_visibles():
    try:
        datos = request.get_json()
        if not datos or not isinstance(datos, list) or len(datos) == 0:
            return jsonify({"error": "No se recibieron datos válidos"}), 400

        df = pd.DataFrame(datos)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Reportes')

            # Aplicar estilos a la hoja
            workbook = writer.book
            hoja = writer.sheets['Reportes']

            # Estilo para títulos
            font = Font(bold=True, size=12, color="000000")  # Letra blanca
            fill = PatternFill(start_color="f5fc19", end_color="f5fc19", fill_type="solid")  # Amarillo

            for col_num, column_title in enumerate(df.columns, 1):
                celda = hoja[f"{get_column_letter(col_num)}1"]
                celda.font = font
                celda.fill = fill
                hoja.column_dimensions[get_column_letter(col_num)].width = max(15, len(column_title) + 5)

        output.seek(0)
        return send_file(
            output,
            download_name="Solicitudes_Visibles.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("Error al exportar Excel:", str(e))
        return jsonify({"error": "Error interno", "details": str(e)}), 500

# Función que se ejecutará programadamente
def actualizar_vacaciones_automaticamente():
    with app.app_context():  # Necesario para acceder al contexto de Flask
        print(f"[{datetime.now()}] Ejecutando actualización de vacaciones...")
        try:
            # Llamar al endpoint como si fuera una petición HTTP
            client = app.test_client()
            response = client.post('/vacaciones_por_ley')
            print("Resultado:", response.get_json())
        except Exception as e:
            print(f"Error al actualizar vacaciones: {e}")

# Configurar el scheduler
scheduler = BackgroundScheduler()
scheduler.add_job(
    actualizar_vacaciones_automaticamente,
    trigger='cron',  # Ejecución programada
    hour=9,          # A las 8:00 AM (ajusta la hora según necesites)
    minute=5,
)
scheduler.start()

# Detener el scheduler al cerrar la aplicación
import atexit
atexit.register(lambda: scheduler.shutdown())



@app.route('/<path:filename>')
def serve_file(filename):
    return send_from_directory('.', filename)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)

